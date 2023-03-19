import openpyxl

# Excelファイル名とシート名を定義する
filename = "example.xlsx"
sheetname = "Sheet1"

# Excelファイルを新規作成し、シートを追加する
wb = openpyxl.Workbook()
ws = wb.active
ws.title = sheetname

# ヘッダーを書き込む
headers = ["Header 1", "Header 2", "Header 3"]
for col, header in enumerate(headers, start=1):
    ws.cell(row=1, column=col, value=header)

# データを書き込む
data = [
    ["Data 1-1", "Data 1-2", "Data 1-3"],
    ["Data 2-1", "Data 2-2", "Data 2-3"],
    ["Data 3-1", "Data 3-2", "Data 3-3"]
]
for row, row_data in enumerate(data, start=2):
    for col, cell_data in enumerate(row_data, start=1):
        ws.cell(row=row, column=col, value=cell_data)

# Excelファイルを保存する
wb.save(filename)
